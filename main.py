import time
import flet as ft
from assets import database as db
import datetime
import openpyxl as op
from openpyxl.styles import Alignment, Font
import requests as r
from threading import Thread
from flet import TextField

items = []
view_table_tables = ""
view_table_date = ""
service_number = len([x for x in db.fetch_data_time() if "PRICE" in x])
price_number = len([x for x in db.fetch_data_time() if "SERVICE" in x])
error = "This field can not be empty"
credentials = {}
done = False


def main(page: ft.Page):
    global credentials
    page.title = "hello world"
    page.window.min_width = 1137
    page.window.min_height = 600
    page.spacing = 0
    page.padding = ft.padding.all(0)
    accent_color = ft.Colors.LIGHT_BLUE_ACCENT_100
    service_number = len([x for x in db.fetch_data_time() if "PRICE" in x])
    price_number = len([x for x in db.fetch_data_time() if "SERVICE" in x])

    # error = "This field can not be empty"
    # credentials = {}
    # done = False

    def regenerate():
        global service_number, price_number
        service_text.value = f"Service table : {len([x for x in db.fetch_data_time() if "SERVICE" in x])}"
        price_text.value = f"Price table :{len([x for x in db.fetch_data_time() if "PRICE" in x])}"
        page.update()

    def uploading():
        upload.visible = True
        page.update()
        time.sleep(5)
        upload.visible = False
        page.update()

    def commenting(_):
        receive()
        page.open(comments)

    page.theme = ft.Theme(
        scrollbar_theme=ft.ScrollbarTheme(
            track_color={
                ft.ControlState.DEFAULT: ft.Colors.TRANSPARENT,
            },
            thickness=5,
            radius=15,
        ),
    )

    class View(ft.Container):
        def __init__(self, date, table):
            super().__init__()
            self.date = date
            self.table = table
            self.ink = True
            self.border = ft.border.all(2, accent_color)
            self.border_radius = 10
            self.ink_color = accent_color
            self.view_btn = ft.IconButton(icon=ft.Icons.VIEW_LIST,
                                          on_click=self.show)
            self.content = (
                ft.Column(
                    controls=[
                        ft.Container(alignment=ft.alignment.bottom_center,
                                     content=ft.Text(str(table))),
                        ft.Container(alignment=ft.alignment.bottom_center,
                                     content=ft.Text(str(self.date)), expand=90),
                        self.view_btn
                    ]
                )
            )

        def show(self, e):
            global view_table_tables, view_table_date
            num = 1
            if self.table == "PRICE":
                view_table_tables = "PRICE"
                view_table_date = str(self.date)
                view_table.rows = []
                items_view = db.fetch_data_saved()
                view_table.columns = [
                    ft.DataColumn(ft.Text("", expand=100, )),
                    ft.DataColumn(ft.Text("Name", expand=100, )),
                    ft.DataColumn(ft.Text("Quantity ", expand=100, )),
                    ft.DataColumn(ft.Text("Category", expand=100, )),
                    ft.DataColumn(ft.Text("Price", expand=100, )),
                    ft.DataColumn(ft.Text("time", expand=100, )),
                ]
                for x in items_view:
                    if self.date in x:
                        name = x[0]
                        quantity = x[1]
                        price_ = x[2]
                        category = x[3]
                        time__ = x[4]
                        view_table.rows.append(ft.DataRow(
                            cells=[
                                ft.DataCell(ft.Text(str(num))),
                                ft.DataCell(ft.Text(name)),
                                ft.DataCell(ft.Text(quantity)),
                                ft.DataCell(ft.Text(category)),
                                ft.DataCell(ft.Text(price_)),
                                ft.DataCell(ft.Text(time__)),
                            ]
                        ))
                        num += 1
                price_table.visible = False
                service_table.visible = False
                view_table.visible = True
                view_container.visible = False
                view_table.update()
                body_container.update()
                page.update()
            elif self.table == "SERVICE":
                view_table_tables = "SERVICE"
                view_table_date = str(self.date)
                view_table.rows = []
                items_view = db.fetch_data_service_saved()
                view_table.columns = [
                    ft.DataColumn(ft.Text("", expand=100, )),
                    ft.DataColumn(ft.Text("Name", expand=100, )),
                    ft.DataColumn(ft.Text("Status", expand=100, )),
                    ft.DataColumn(ft.Text("Category", expand=100, )),
                    ft.DataColumn(ft.Text("Price", expand=100, )),
                    ft.DataColumn(ft.Text("time", expand=100, )),
                ]
                for x in items_view:
                    if self.date in x:
                        name = x[0]
                        status = x[1]
                        category = x[2]
                        price_ = x[3]
                        time_ = x[4]
                        view_table.rows.append(ft.DataRow(
                            cells=[
                                ft.DataCell(ft.Text(str(num))),
                                ft.DataCell(ft.Text(name)),
                                ft.DataCell(ft.Text(status)),
                                ft.DataCell(ft.Text(category)),
                                ft.DataCell(ft.Text(price_)),
                                ft.DataCell(ft.Text(time_)),
                            ]
                        ))
                        num += 1
                price_table.visible = False
                service_table.visible = False
                view_table.visible = True
                view_container.visible = False
                view_table.update()
                body_container.update()
                page.update()

    def stat_edit(e):
        new_ = e.control.value
        data = e.control.data
        db.update_item_service(data[0], new_, data[2], data[3], data[4])
        service(e)
        service_table.update()
        page.update()

    def pric_edit(e):
        new__ = e.control.value
        data = e.control.data
        db.update_item_service(data[0], data[1], data[2], new__, data[4])
        service(e)
        service_table.update()
        page.update()

    def remove(e):
        data = e.control.data
        db.remove_item_service(data[0])
        service(e)
        service_table.update()
        page.update()

    def quantity_edit(e):
        new___ = e.control.value
        data = e.control.data
        db.update_item(data[0], new___, data[2], data[3], data[4])
        price(e)
        service_table.update()
        page.update()

    def pric_edit_service(e):
        new1 = e.control.value
        data = e.control.data
        db.update_item(data[0], data[1], data[2], new1, data[4])
        price(e)
        service_table.update()
        page.update()

    def remove_service(e):
        data = e.control.data
        db.remove_item(data[0])
        price(e)
        service_table.update()
        page.update()

    def append_service():
        num = 1
        for x in db.fetch_data_service():
            name = x[0]
            status = x[1]
            category = x[2]
            price_ = x[3]
            time_ = x[4]
            service_table.rows.append(ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(str(num))),
                    ft.DataCell(ft.Text(name)),
                    ft.DataCell(ft.TextField(max_length=7, value=status, expand=2, on_submit=stat_edit, data=x)),
                    ft.DataCell(ft.Text(category)),
                    ft.DataCell(
                        ft.TextField(max_length=10, value=price_, expand=2, on_submit=pric_edit, data=x)),
                    ft.DataCell(ft.Text(time_)),
                    ft.DataCell(
                        ft.IconButton(icon=ft.Icons.DELETE, data=x, on_click=remove))

                ]
            ))
            num += 1

    def append_price():
        num = 1
        for x in db.fetch_data():
            name = x[0]
            quantity = x[1]
            category = x[2]
            price_ = x[3]
            time_ = x[4]
            price_table.rows.append(ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(str(num))),
                    ft.DataCell(ft.Text(name)),
                    ft.DataCell(
                        ft.TextField(max_length=10, value=quantity, expand=2, on_submit=quantity_edit, data=x), ),
                    ft.DataCell(ft.Text(category)),
                    ft.DataCell(
                        ft.TextField(max_length=10, value=price_, expand=2, on_submit=pric_edit_service, data=x)),
                    ft.DataCell(ft.Text(time_)),
                    ft.DataCell(content=ft.IconButton(icon=ft.Icons.DELETE, data=x, on_click=remove_service))]
            ))
            num += 1

    def export_to_file(e):
        page.close(price_export_dates)
        date = dates.value
        file_data = db.fetch_data_saved()
        new_file_data = []
        for x in file_data:
            if date in x:
                new_file_data.append(x)

        num = 1
        row = 5
        column = 1
        company_name__ = "Company Name"
        workbook = op.Workbook()

        sheet = workbook.active

        sheet.merge_cells("A1:E1")
        sheet.merge_cells("A2:E2")
        sheet.merge_cells("A3:E3")

        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet["A2"].alignment = Alignment(horizontal="center")
        sheet["A3"].alignment = Alignment(horizontal="center")

        sheet.column_dimensions["A"].width = 5
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 20
        sheet.row_dimensions[1].height = 30

        font = Font(size=28)
        font1 = Font(size=18)

        name = sheet.cell(row=1, column=1)
        name.font = font
        name.value = company_name__
        name = sheet.cell(row=2, column=1)
        name.font = font1
        name.value = "Price Stock Table"
        name = sheet.cell(row=3, column=1)
        name.font = font1
        name.value = f"Date : {date}"

        number = sheet.cell(row=4, column=1)
        number.value = "#"

        name = sheet.cell(row=4, column=2)
        name.value = "Name"

        quantity = sheet.cell(row=4, column=3)
        quantity.value = "Quantity"

        price_file = sheet.cell(row=4, column=4)
        price_file.value = "Price"

        category = sheet.cell(row=4, column=5)
        category.value = "Category"

        for x in new_file_data:
            sheet.cell(row=row, column=1).value = num
            sheet.cell(row=row, column=2).value = x[0]
            sheet.cell(row=row, column=3).value = x[1]
            sheet.cell(row=row, column=4).value = x[2]
            sheet.cell(row=row, column=5).value = x[3]
            row += 1
            num += 1

        workbook.save(f"{date} Price Stock Table.xlsx")

    def export_to_file_service(e):
        page.close(service_export_dates)
        date = dates.value
        file_data = db.fetch_data_service_saved()
        new_file_data = []
        for x in file_data:
            if date in x:
                new_file_data.append(x)

        num = 1
        row = 5
        column = 1
        company_name1 = "Company Name"
        workbook = op.Workbook()

        sheet = workbook.active

        sheet.merge_cells("A1:E1")
        sheet.merge_cells("A2:E2")
        sheet.merge_cells("A3:E3")

        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet["A2"].alignment = Alignment(horizontal="center")
        sheet["A3"].alignment = Alignment(horizontal="center")

        sheet.column_dimensions["A"].width = 5
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 20
        sheet.row_dimensions[1].height = 30

        font = Font(size=28)
        font1 = Font(size=18)

        name = sheet.cell(row=1, column=1)
        name.font = font
        name.value = company_name1
        name = sheet.cell(row=2, column=1)
        name.font = font1
        name.value = "Service Stock Table"
        name = sheet.cell(row=3, column=1)
        name.font = font1
        name.value = f"Date : {date}"

        number = sheet.cell(row=4, column=1)
        number.value = "#"

        name = sheet.cell(row=4, column=2)
        name.value = "Name"

        quantity = sheet.cell(row=4, column=3)
        quantity.value = "Quantity"

        price_file = sheet.cell(row=4, column=4)
        price_file.value = "Price"

        category = sheet.cell(row=4, column=5)
        category.value = "Category"

        for x in new_file_data:
            sheet.cell(row=row, column=1).value = num
            sheet.cell(row=row, column=2).value = x[0]
            sheet.cell(row=row, column=3).value = x[1]
            sheet.cell(row=row, column=4).value = x[2]
            sheet.cell(row=row, column=5).value = x[3]
            row += 1
            num += 1

        workbook.save(f"{date} Service Stock Table.xlsx")

    def filter_table(_):
        global view_table_tables, view_table_date
        service_table.rows = []
        price_table.rows = []
        searching = search_field.value

        if searching == "":
            if service_table.visible:
                append_service()
            elif price_table.visible:
                append_price()
            elif view_table.visible:
                num = 1
                if view_table_tables == "PRICE":
                    view_table_tables = "PRICE"
                    view_table_date = str(view_table_date)
                    view_table.rows = []
                    items_view = db.fetch_data_saved()
                    view_table.columns = [
                        ft.DataColumn(ft.Text("", expand=100, )),
                        ft.DataColumn(ft.Text("Name", expand=100, )),
                        ft.DataColumn(ft.Text("Quantity ", expand=100, )),
                        ft.DataColumn(ft.Text("Price", expand=100, )),
                        ft.DataColumn(ft.Text("Category", expand=100, )),
                        ft.DataColumn(ft.Text("time", expand=100, )),
                    ]
                    for x in items_view:
                        if view_table_date in x:
                            name = x[0]
                            quantity = x[1]
                            price_ = x[2]
                            category = x[3]
                            time1 = x[4]
                            view_table.rows.append(ft.DataRow(
                                cells=[
                                    ft.DataCell(ft.Text(str(num))),
                                    ft.DataCell(ft.Text(name)),
                                    ft.DataCell(ft.TextField(quantity)),
                                    ft.DataCell(ft.Text(price_)),
                                    ft.DataCell(ft.Text(category)),
                                    ft.DataCell(ft.Text(time1)),
                                ]
                            ))
                            num += 1
                    price_table.visible = False
                    service_table.visible = False
                    view_table.visible = True
                    view_container.visible = False
                    view_table.update()
                    body_container.update()
                    page.update()
                elif view_table_tables == "SERVICE":
                    view_table_tables = "SERVICE"
                    view_table_date = str(view_table_date)
                    view_table.rows = []
                    items_view = db.fetch_data_service_saved()
                    view_table.columns = [
                        ft.DataColumn(ft.Text("", expand=100, )),
                        ft.DataColumn(ft.Text("Name", expand=100, )),
                        ft.DataColumn(ft.Text("Quantity", expand=100, )),
                        ft.DataColumn(ft.Text("Service", expand=100, )),
                        ft.DataColumn(ft.Text("Category", expand=100, )),
                        ft.DataColumn(ft.Text("time", expand=100, )),
                    ]
                    for x in items_view:
                        if view_table_date in x:
                            name = x[0]
                            quantity = x[1]
                            status = x[2]
                            category = x[3]
                            time = x[4]
                            view_table.rows.append(ft.DataRow(
                                cells=[
                                    ft.DataCell(ft.Text(str(num))),
                                    ft.DataCell(ft.Text(name)),
                                    ft.DataCell(ft.Text(quantity)),
                                    ft.DataCell(ft.Text(status)),
                                    ft.DataCell(ft.Text(category)),
                                    ft.DataCell(ft.Text(time)),
                                ]
                            ))
                            num += 1
                    price_table.visible = False
                    service_table.visible = False
                    view_table.visible = True
                    view_container.visible = False
                    view_table.update()
                    body_container.update()
                    page.update()
        else:
            service_searched_data = db.fetch_data_service()
            if service_table.visible:
                num = 1
                for x in service_searched_data:

                    if searching in x[0]:
                        name = x[0]
                        status = x[1]
                        category = x[2]
                        price_ = x[3]
                        time_ = x[4]
                        service_table.rows.append(ft.DataRow(
                            cells=[
                                ft.DataCell(ft.Text(str(num))),
                                ft.DataCell(ft.Text(name)),
                                ft.DataCell(ft.TextField(value=status, expand=2, on_submit=stat_edit, data=x)),
                                ft.DataCell(ft.Text(category)),
                                ft.DataCell(
                                    ft.TextField(value=price_, expand=2, on_submit=pric_edit, data=x)),
                                ft.DataCell(ft.Text(time_)),
                                ft.DataCell(
                                    ft.IconButton(icon=ft.Icons.DELETE, data=x, on_click=remove))

                            ]
                        ))
                        num += 1

            elif price_table.visible:
                price_searched_data = db.fetch_data()
                num = 1
                for y in price_searched_data:

                    if searching in y[0]:
                        name = y[0]
                        quantity = y[1]
                        category = y[2]
                        price_ = y[3]
                        time_ = y[4]
                        price_table.rows.append(ft.DataRow(
                            cells=[
                                ft.DataCell(ft.Text(str(num))),
                                ft.DataCell(ft.Text(name)),
                                ft.DataCell(ft.TextField(value=quantity, expand=2, on_submit=quantity_edit, data=y), ),
                                ft.DataCell(ft.Text(category)),
                                ft.DataCell(ft.TextField(value=price_, expand=2, on_submit=pric_edit_service, data=y)),
                                ft.DataCell(ft.Text(time_)),
                                ft.DataCell(
                                    content=ft.IconButton(icon=ft.Icons.DELETE, data=y, on_click=remove_service))

                            ]
                        ))
                        num += 1

            elif view_table.visible:
                num = 1
                if view_table_tables == "PRICE":
                    view_table.rows = []
                    items_view = db.fetch_data_saved()
                    view_table.columns = [
                        ft.DataColumn(ft.Text("", expand=100, )),
                        ft.DataColumn(ft.Text("Name", expand=100, )),
                        ft.DataColumn(ft.Text("Quantity ", expand=100, )),
                        ft.DataColumn(ft.Text("Category", expand=100, )),
                        ft.DataColumn(ft.Text("Price", expand=100, )),
                        ft.DataColumn(ft.Text("time", expand=100, )),
                    ]
                    for x in items_view:
                        if view_table_date in x:
                            if searching in x[0]:
                                name = x[0]
                                quantity = x[1]
                                category = x[2]
                                price_ = x[3]
                                time = x[4]
                                view_table.rows.append(ft.DataRow(
                                    cells=[
                                        ft.DataCell(ft.Text(str(num))),
                                        ft.DataCell(ft.Text(name)),
                                        ft.DataCell(ft.Text(quantity)),
                                        ft.DataCell(ft.Text(price_)),
                                        ft.DataCell(ft.Text(category)),
                                        ft.DataCell(ft.Text(time)),
                                    ]
                                ))
                                num += 1
                    price_table.visible = False
                    service_table.visible = False
                    view_table.visible = True
                    view_container.visible = False
                    view_table.update()
                    body_container.update()
                    page.update()
                elif view_table_tables == "SERVICE":
                    view_table_tables = "SERVICE"
                    view_table_date = str(view_table_date)
                    view_table.rows = []
                    items_view = db.fetch_data_service_saved()
                    view_table.columns = [
                        ft.DataColumn(ft.Text("", expand=100, )),
                        ft.DataColumn(ft.Text("Name", expand=100, )),
                        ft.DataColumn(ft.Text("Status", expand=100, )),
                        ft.DataColumn(ft.Text("Category", expand=100, )),
                        ft.DataColumn(ft.Text("Price", expand=100, )),
                        ft.DataColumn(ft.Text("time", expand=100, )),
                    ]
                    for x in items_view:
                        if view_table_date in x:
                            if searching in x[0]:
                                name = x[0]
                                status = x[1]
                                category = x[2]
                                price_ = x[3]
                                time2 = x[4]
                                view_table.rows.append(ft.DataRow(
                                    cells=[
                                        ft.DataCell(ft.Text(str(num))),
                                        ft.DataCell(ft.Text(name)),
                                        ft.DataCell(ft.Text(status)),
                                        ft.DataCell(ft.Text(category)),
                                        ft.DataCell(ft.Text(price_)),
                                        ft.DataCell(ft.Text(time2)),
                                    ]
                                ))
                                num += 1
                    price_table.visible = False
                    service_table.visible = False
                    view_table.visible = True
                    view_container.visible = False
                    view_table.update()
                    body_container.update()
                    page.update()
        service_table.update()
        price_table.update()
        page.update()

    bg_img = ft.Container(
        ft.Image(src="assets/vecteezy_abstract-background-design-background-texture-design-with_18752866-1.jpg",
                 opacity=0.5, scale=2))

    def login_user():
        title_container.visible = False
        body_container.visible = False
        bottom_container.visible = False
        bg_img.visible = True
        try:
            read()
            page.open(login)
            company_name_title.value = credentials["company_name"]
            page.update()

        except IndexError:
            page.open(authenticate)

    def read():
        global credentials
        datum = db.fetch_data_user()
        credentials = dict(company_name=datum[0][0], password=datum[0][1], email=datum[0][2],
                           application_password=datum[0][3])

    def login_user_wt_pass(e):
        bg_img.visible = False
        if password.value == "":
            password.error_text = "empty field"
            page.update()
        elif password.value != credentials["password"]:
            password.error_text = "wrong password"
            password.update()
            page.update()
        elif password.value == credentials["password"]:
            page.close(login)
            Thread(target=old_upload).start()
            Thread(target=old_upload_service).start()
            title_container.visible = True
            body_container.visible = True
            bottom_container.visible = True

            page.update()

    def changing(_):
        if old.value == credentials["password"]:
            if new.value != "" or confirm.value != "":
                if new.value == confirm.value:
                    db.update_item_user(credentials["company_name"],
                                        new.value,
                                        credentials["email"],
                                        confirm.value)
                    read()
                    page.close(settings)
                    new.value = ""
                    old.value = ""
                    confirm.value = ""
                    page.update()
                else:
                    new.error_text = "password does no match"
                    confirm.error_text = "password does no match"
                    page.update()
            else:
                new.error_text = "check details"
                confirm.error_text = "check details"
                page.update()
        else:
            old.error_text = "wrong password"
            page.update()

    def login_btn1(e):
        connectionerror.visible = False
        page.update()
        if company_name.value == "":
            company_name.error_text = error
            email.error_text = None
            password.error_text = None
            authenticate.update()
        else:
            company_name.error_text = None
            if email.value == "":
                email.error_text = error
                password.error_text = None
                authenticate.update()
            else:
                email.error_text = None
                if password.value == "":
                    password.error_text = error
                else:
                    bg_img.visible = False
                    try:
                        d = r.post(url="http://royal-blueocelot.onpella.app/verify",
                                   json=dict(company_name=company_name.value,
                                             email=email.value,
                                             password=password.value))
                        if d.json() == "verified":
                            db.add_item_user(company_name.value, password.value, email.value, password.value)
                            page.close(authenticate)
                            title_container.visible = True
                            body_container.visible = True
                            bottom_container.visible = True
                            read()
                            company_name_title.value = credentials["company_name"]
                            page.update()
                            Thread(target=old_upload).start()
                            Thread(target=old_upload_service).start()
                        else:
                            company_name.error_text = "check your details"
                            password.error_text = "check your details"
                            email.error_text = "check your details"
                            page.update()
                    except r.ConnectionError:
                        connectionerror.visible = True
                        page.update()

    company_name = TextField(helper_text="Company Name", border_color="blue", border_radius=10, )
    email = TextField(helper_text="Email", border_color="blue", border_radius=10, )
    password = TextField(helper_text="Password", border_color="blue", border_radius=10, password=True, color="white",
                         can_reveal_password=True)

    view_table = ft.DataTable(
        column_spacing=15,

        expand=100,
        visible=False,
        columns=[
            ft.DataColumn(ft.Text("", expand=100, )),
            ft.DataColumn(ft.Text("Name", expand=100, )),
            ft.DataColumn(ft.Text("Quantity", expand=100, )),
            ft.DataColumn(ft.Text("Category", expand=100, )),
            ft.DataColumn(ft.Text("price", expand=100, )),
            ft.DataColumn(ft.Text("time", expand=100, )),
        ],
        rows=[
        ],
    )

    fill_containers = ft.Container(expand=1)
    fill_containers_ = ft.Container(expand=1)

    price_table = ft.DataTable(
        column_spacing=15,
        sort_column_index=0,
        sort_ascending=True,
        show_checkbox_column=True,
        visible=False,
        expand=8,
        columns=[
            ft.DataColumn(ft.Text("")),
            ft.DataColumn(ft.Text("Name")),
            ft.DataColumn(ft.Text("Quantity")),
            ft.DataColumn(ft.Text("Category")),
            ft.DataColumn(ft.Text("Price")),
            ft.DataColumn(ft.Text("Date")),
            ft.DataColumn(ft.Text("Actions")),
        ],
        rows=[
        ],
    )

    service_table = ft.DataTable(
        column_spacing=15,
        expand=8,
        sort_column_index=0,
        sort_ascending=True,
        show_checkbox_column=True,
        visible=False,
        columns=[
            ft.DataColumn(ft.Text("", expand=100, )),
            ft.DataColumn(ft.Text("Name", expand=100, )),
            ft.DataColumn(ft.Text("Status", expand=100, )),
            ft.DataColumn(ft.Text("Category", expand=100, )),
            ft.DataColumn(ft.Text("Price", expand=100, )),
            ft.DataColumn(ft.Text("Date", expand=100, )),
            ft.DataColumn(ft.Text("Actions", expand=100, )),
        ],
        rows=[
        ],
    )

    search_field = ft.TextField(hint_text="search", border_radius=30, height=40, width=300,
                                enable_suggestions=True,
                                dense=True, on_change=filter_table, icon=ft.Icons.SEARCH)

    # read()
    company_name_title = ft.Text(value=f"", size=25)

    title_container = ft.Container(
        border=ft.border.all(2, accent_color),
        height=20,
        expand=10,
        visible=False,
        content=(
            ft.Row(
                [
                    ft.Container(
                        alignment=ft.alignment.center_left,
                        expand=60,
                        padding=ft.padding.only(left=50),
                        border_radius=10,

                        content=(
                            company_name_title
                        ),
                    ),
                    ft.Container(
                        alignment=ft.alignment.center,
                        expand=30,
                        border_radius=10,
                        content=(
                            ft.Row(
                                controls=[
                                    search_field
                                ]
                            )
                        )
                    ),
                ]
            )
        )
    )

    def date_update():
        date_ = datetime.date.today()
        db.update_date(date_)
        db.update_date_service(date_)

    def view(e):
        global items
        search_field.value = ""
        times = db.fetch_data_time()
        items = []
        if (len(times)) == 0:
            view_container.visible = False
            empty_text.visible = True
        else:
            for x in times:
                items.append(View(x[0], x[1]))
            view_container.visible = True
            empty_text.visible = False
        view_container.controls = items
        price_table.visible = False
        service_table.visible = False
        view_table.visible = False
        view_container.update()
        body_container.update()
        page.update()

    def price_data(e):
        if name_.value != "":
            if category_.value != "":
                if quantity_.value.isdigit():
                    if price__.value.isdigit():
                        search_field.value = ""
                        search_field.update()
                        name = name_.value
                        quantity = quantity_.value
                        price_ = price__.value
                        category = category_.value
                        time3 = ""
                        price_table.rows = []
                        db.add_item(name, quantity, category, price_, time3)
                        date_update()
                        append_price()
                        name_.value = ""
                        category_.value = ""
                        price__.value = ""
                        quantity_.value = ""
                        add_item.update()
                        view_container.visible = False
                        empty_text.visible = False
                        view_table.visible = False
                        price_table.visible = True
                        service_table.visible = False
                        body_container.update()
                    else:
                        page.update()
                        price__.error_text, quantity_.error_text, name_.error_text, category_.error_text \
                            = ("check your input", "check your input", "check your input",
                               "check your input")
                else:
                    page.update()
                    price__.error_text, quantity_.error_text, name_.error_text, category_.error_text \
                        = ("check your input", "check your input", "check your input",
                           "check your input")
            else:
                page.update()
                price__.error_text, quantity_.error_text, name_.error_text, category_.error_text \
                    = ("check your input", "check your input", "check your input",
                       "check your input")
        else:
            page.update()
            price__.error_text, quantity_.error_text, name_.error_text, category_.error_text \
                = ("check your input", "check your input", "check your input",
                   "check your input")

    def service_data(e):
        if name__.value != "":
            if status__.value != "":
                if category__.value != "":
                    if price___.value.isdigit():
                        search_field.value = ""
                        search_field.update()
                        name = name__.value
                        status = status__.value
                        category = category__.value
                        price_ = price___.value
                        time4 = ""
                        service_table.rows = []
                        db.add_item_service(name, status, category, price_, time4)
                        date_update()
                        append_service()
                        name__.value = ""
                        status__.value = ""
                        category__.value = ""
                        price___.value = ""
                        service_item.update()
                        view_container.visible = False
                        empty_text.visible = False
                        view_table.visible = False
                        price_table.visible = False
                        service_table.visible = True
                        body_container.update()
                    else:
                        page.update()
                        price___.error_text, status__.error_text, name__.error_text, category__.error_text \
                            = ("check your input", "check your input", "check your input",
                               "check your input")
                else:
                    page.update()
                    price___.error_text, status__.error_text, name__.error_text, category__.error_text \
                        = ("check your input", "check your input", "check your input",
                           "check your input")
            else:
                page.update()
                price___.error_text, status__.error_text, name__.error_text, category__.error_text \
                    = ("check your input", "check your input", "check your input",
                       "check your input")
        else:
            page.update()
            price___.error_text, status__.error_text, name__.error_text, category__.error_text \
                = ("check your input", "check your input", "check your input",
                   "check your input")

    def price(e):
        search_field.value = ""
        search_field.update()
        price_table.rows = []
        date_update()
        append_price()
        view_container.visible = False
        empty_text.visible = False
        view_table.visible = False
        price_table.visible = True
        service_table.visible = False
        body_container.update()

    def service(e):
        search_field.value = ""
        search_field.update()
        service_table.rows = []
        date_update()
        append_service()
        view_container.visible = False
        view_table.visible = False
        empty_text.visible = False
        price_table.visible = False
        service_table.visible = True
        body_container.update()

    def receive():
        print(credentials)
        email = credentials["email"]
        try:
            d = r.post(url="http://royal-blueocelot.onpella.app/comments", json=dict(email=email))
            comment = []
            rec = d.json()
            for x in rec[1]:
                comment.append(ft.Text(x))
            comment_data.controls = comment
            comments.update()
            page.update()
            time.sleep(1)
        except r.ConnectionError:
            pass

    def old_upload():
        print("old_upload")
        num = 0
        try:
            data_item = db.fetch_data_items_uploaded()
            data = db.fetch_data()
            new_data = []
            read()
            for x in data_item:
                x = x[0]
                for y in data:
                    print(x in y)
                    if x in y:
                        new_data.append(y)
                r.post(url="http://royal-blueocelot.onpella.app/price_url",
                       json=dict(email=credentials["email"], data=new_data))
                new_data = []
                num += 1
        except r.ConnectionError:
            for x in range(num - 1):
                db.remove_item_upload(x)

    def old_upload_service():
        print("old_upload_service")
        num = 0
        try:
            data_item = db.fetch_data_items_uploaded_service()
            data = db.fetch_data_service()
            new_data = []
            read()
            for x in data_item:
                x = x[0]
                for y in data:
                    if x in y:
                        new_data.append(y)
                r.post(url="http://royal-blueocelot.onpella.app/service_url",
                       json=dict(email=credentials["email"], data=new_data))
                new_data = []
                num += 1
        except r.ConnectionError:
            for x in range(num - 1):
                db.remove_item_upload_service(x)

    def price_upload():
        try:
            data = db.fetch_data()
            read()
            r.post(url="http://royal-blueocelot.onpella.app/price_url",
                   json=dict(email=credentials["email"], data=data))
        except r.ConnectionError:
            data = db.fetch_data()
            db.add_item_uploaded(data[0][-1])

    def save_price(e):
        date_update()
        price_text.value = f"Price table : {len([x for x in db.fetch_data_time() if "PRICE" in x])}"
        service_text.value = f"Service table : {len([x for x in db.fetch_data_time() if "SERVICE" in x])}"
        page.update()
        bottom_container.update()
        if str(db.check()) == "0":
            page.open(no_data)
            page.update()
        else:
            result = db.save()
            if result:
                page.open(item_exist)
            else:
                db.add_item_time(str(datetime.date.today()), "PRICE")
                page.open(success)
                Thread(target=price_upload).start()
                uploading()

        regenerate()

    def service_upload():
        try:
            data = db.fetch_data_service()
            read()
            r.post(url="http://royal-blueocelot.onpella.app/service_url",
                   json=dict(email=credentials["email"], data=data))
        except r.ConnectionError:
            data = db.fetch_data_service()
            db.add_item_uploaded_service(data[0][-1])

    def save_service(e):
        date_update()
        if str(db.check_service()) == "0":
            page.open(no_data)
            page.update()
        else:
            result = db.save_service()
            if result:
                page.open(item_exist)
            else:
                db.add_item_time(str(datetime.date.today()), "SERVICE")
                page.open(success)
                Thread(target=service_upload).start()
                uploading()

        regenerate()

    view_container = ft.GridView(
        auto_scroll=True,
        runs_count=8,
        padding=20,
        visible=False,

    )

    item_exist = ft.AlertDialog(
        modal=True,
        title=ft.Text("Stock Exist", size=30, text_align=ft.alignment.center),
        alignment=ft.alignment.center,
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        content=(
            ft.Container(
                content=ft.Text("You have already saved a stock with this date")
            )
        ),
        actions=[
            ft.TextButton("Ok", on_click=lambda e: page.close(item_exist)),
        ],
        actions_alignment=ft.MainAxisAlignment.SPACE_AROUND,
    )

    no_data = ft.AlertDialog(
        modal=True,
        title=ft.Text("Empty Data", size=30, text_align=ft.alignment.center),
        alignment=ft.alignment.center,
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        content=(
            ft.Container(
                content=ft.Text("No data found in the table")
            )
        ),
        actions=[
            ft.TextButton("Ok", on_click=lambda e: page.close(no_data)),
        ],
        actions_alignment=ft.MainAxisAlignment.SPACE_AROUND,
    )

    success = ft.AlertDialog(
        modal=True,
        title=ft.Text("Success", text_align=ft.alignment.center, size=30),
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        content=(
            ft.Container(
                content=ft.Text("Your stock has been successfully save", text_align=ft.alignment.center)
            )
        ),
        actions=[
            ft.TextButton("Ok", on_click=lambda e: page.close(success)),
        ],
        actions_alignment=ft.MainAxisAlignment.SPACE_AROUND,
    )

    dates = ft.TextField(helper_text="Enter date")

    price_export_dates = ft.AlertDialog(
        modal=True,
        title=ft.Text("Enter Date", text_align=ft.alignment.center, size=30),
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        content=(
            ft.Container(
                content=dates
            )
        ),
        actions=[
            ft.TextButton("Ok", on_click=lambda e: export_to_file(e)),
        ],
        actions_alignment=ft.MainAxisAlignment.SPACE_AROUND,
    )

    service_export_dates = ft.AlertDialog(
        modal=True,
        title=ft.Text("Enter Date", text_align=ft.alignment.center, size=30),
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        content=(
            ft.Container(
                content=dates
            )
        ),
        actions=[
            ft.TextButton("Ok", on_click=lambda e: export_to_file_service(e)),
        ],
        actions_alignment=ft.MainAxisAlignment.SPACE_AROUND,
    )
    name_ = ft.TextField(max_length=80, helper_text="Name")
    quantity_ = ft.TextField(max_length=10, helper_text="Quantity")
    price__ = ft.TextField(max_length=10, helper_text="Price")
    category_ = ft.TextField(max_length=30, helper_text="Category")
    add_item = ft.AlertDialog(
        modal=True,
        title=ft.Text("Please confirm"),
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        content=(
            ft.Column(
                controls=[name_,
                          quantity_,
                          price__,
                          category_]
            )
        ),
        actions=[
            ft.TextButton("Add", on_click=price_data),
            ft.TextButton("Back", on_click=lambda e: page.close(add_item)),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )
    name__ = ft.TextField(max_length=80, helper_text="Name")
    price___ = ft.TextField(max_length=10, helper_text="Price")
    category__ = ft.TextField(max_length=30, helper_text="Category")
    status__ = ft.Dropdown(label="Status",
                           options=[
                               ft.dropdown.Option("Working"),
                               ft.dropdown.Option("Faulty"),
                           ], )
    service_item = ft.AlertDialog(
        modal=True,
        title=ft.Text("Please confirm"),
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        content=(
            ft.Column(
                controls=[name__,
                          status__,
                          category__,
                          price___
                          ]
            )
        ),
        actions=[
            ft.TextButton("Add", on_click=service_data),
            ft.TextButton("Back", on_click=lambda e: page.close(service_item)),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    notifications = ft.AlertDialog(
        modal=True,
        title=ft.Text("Notifications Will Appear Here"),
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        actions=[
            # ft.TextButton("Yes", on_click=lambda e: page.close(notifications)),
            ft.TextButton("Close", on_click=lambda e: page.close(notifications)),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    show = ft.AlertDialog(
        modal=True,
        title=ft.Text("Show Stock Table"),
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        actions=[
            ft.TextButton("Price", on_click=price),
            ft.TextButton("Service", on_click=service),
            ft.TextButton("Back", on_click=lambda e: page.close(show)),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    add = ft.AlertDialog(
        modal=True,
        title=ft.Text("""     Add New Data 
      To Stock Table
      
      
      """),
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        actions=[
            ft.TextButton("Price", on_click=lambda e: page.open(add_item)),
            ft.TextButton("Service", on_click=lambda e: page.open(service_item)),
            ft.TextButton("Back", on_click=lambda e: page.close(add)),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    export = ft.AlertDialog(
        modal=True,
        title=ft.Text("Export to file"),
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        actions=[
            ft.TextButton("Price", on_click=lambda e: page.open(price_export_dates)),
            ft.TextButton("Service", on_click=lambda e: page.open(service_export_dates)),
            ft.TextButton("Back", on_click=lambda e: page.close(export)),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    save = ft.AlertDialog(
        modal=True,
        title=ft.Text("Save Stock Table"),
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        actions=[
            ft.TextButton("Price", on_click=save_price),
            ft.TextButton("Service", on_click=save_service),
            ft.TextButton("Back", on_click=lambda e: page.close(save)),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    old = ft.TextField(helper_text="Old password", can_reveal_password=True, password=True)
    new = ft.TextField(helper_text="New password", can_reveal_password=True, password=True)
    confirm = ft.TextField(helper_text="Confirm New password", can_reveal_password=True, password=True)
    change = ft.TextButton(text="Change Password", icon=ft.Icons.CHANGE_CIRCLE, on_click=changing)

    settings = ft.AlertDialog(
        modal=True,
        title=ft.Text("Settings"),
        content_padding=30,
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        content=ft.Column(
            controls=[
                ft.Text("Change password", size=25),
                old,
                new,
                confirm,
                change
            ]
        ),
        actions=[
            # ft.TextButton("Yes", on_click=lambda e: set()),
            ft.TextButton("Back", on_click=lambda e: page.close(settings)),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    connectionerror = ft.Text("Connection Error", visible=False, color="red", size=40)

    authenticate = ft.AlertDialog(
        modal=True,
        title=ft.Text("Login"),
        content_padding=30,
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        content=ft.Column(
            controls=[
                company_name,
                email,
                password,
                connectionerror
            ]
        ),
        actions=[
            ft.TextButton("Login", on_click=lambda e: login_btn1(e)),
            ft.TextButton("Exit", on_click=lambda e: page.window.close()),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    login = ft.AlertDialog(
        modal=True,
        title=ft.Text("Login"),
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        content=ft.Column(
            controls=[
                password
            ]
        ),
        actions=[
            ft.TextButton("Login", on_click=lambda e: login_user_wt_pass(e)),
            ft.TextButton("Exit", on_click=lambda e: page.window.close()),
        ],
        actions_alignment=ft.MainAxisAlignment.END
    )
    comment_data = ft.Column(scroll=ft.ScrollMode.AUTO, controls=[])
    comments = ft.AlertDialog(
        modal=True,
        title=ft.Text("Comments"),
        shadow_color=accent_color,
        surface_tint_color=accent_color,
        elevation=10,
        content_padding=10,
        content=comment_data,
        actions=[
            ft.TextButton("Back", on_click=lambda e: page.close(comments)),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    empty_text = ft.Text(value="No items found", size=55, visible=False)

    body_container = ft.Row(
        visible=False,
        controls=
        [

            ft.Container(
                border=ft.border.all(2, accent_color),
                alignment=ft.alignment.center,
                expand=6,
                # border_radius=10,
                padding=ft.padding.only(top=10),
                content=(

                    ft.Column(
                        spacing=20,
                        controls=[
                            ft.TextButton(text="Show", on_click=lambda e: page.open(show)),
                            ft.TextButton(text="Add item", on_click=lambda e: page.open(add)),
                            ft.TextButton(text="view", on_click=view),
                            ft.TextButton(text="save", on_click=lambda e: page.open(save)),
                            ft.TextButton(text="Export", on_click=lambda e: page.open(export)),
                            ft.TextButton(text="exit", on_click=lambda e: page.window.close()),
                        ]
                    )
                )
            ),

            ft.Container(
                border=ft.border.all(2, accent_color),
                alignment=ft.alignment.center,
                expand=75,

                # border_radius=10,
                content=(
                    ft.Column(
                        alignment=ft.MainAxisAlignment.CENTER,
                        controls=[
                            ft.Column(spacing=10,
                                      height=200,
                                      width=float("inf"),
                                      expand=1,
                                      scroll=ft.ScrollMode.ALWAYS,
                                      controls=[view_container, ft.Container(alignment=ft.alignment.center,
                                                                             content=ft.Column(
                                                                                 controls=[ft.Row(
                                                                                     scroll=ft.ScrollMode.ALWAYS,
                                                                                     controls=[fill_containers,
                                                                                               price_table,
                                                                                               fill_containers_]),
                                                                                     ft.Row(
                                                                                         scroll=ft.ScrollMode.ALWAYS,
                                                                                         controls=[fill_containers,
                                                                                                   service_table,
                                                                                                   fill_containers_]),
                                                                                     empty_text,
                                                                                     ft.Row(
                                                                                         scroll=ft.ScrollMode.ALWAYS,
                                                                                         controls=[fill_containers,
                                                                                                   view_table,
                                                                                                   fill_containers_,
                                                                                                   ])]))]),
                        ]
                    )
                ),
            ),

            ft.Container(
                border=ft.border.all(2, accent_color),
                alignment=ft.alignment.center,
                expand=3,
                # border_radius=10,
                content=(
                    ft.Column(
                        controls=[
                            ft.IconButton(on_click=lambda e: commenting(e),
                                          icon=ft.Icons.COMMENT),
                            ft.IconButton(icon=ft.Icons.SETTINGS, on_click=lambda e: page.open(settings)),
                            ft.IconButton(icon=ft.Icons.HELP, url="http://127.0.0.1:5000/login"),
                        ]
                    )
                )
            ),
        ],
        spacing=0,
        expand=85,
    )

    date_update()

    service_text = ft.Text(value=f"Service table : {service_number}")

    price_text = ft.Text(value=f"Price table : {price_number}")

    upload = ft.Container(visible=False, content=ft.Text(value="Uploading ...."))

    bottom_container = ft.Container(
        border=ft.border.all(2, accent_color),
        height=20,
        visible=False,
        expand=5,
        content=ft.Row(
            controls=[
                service_text,
                price_text,
                upload
            ])
    )

    page.add(item_exist, comments, settings, notifications, service_item, add_item, title_container, body_container,
             bottom_container, success, no_data, show, export, save, add, price_export_dates, service_export_dates,
             login, authenticate, bg_img)

    login_user()



if __name__ == "__main__":
    ft.app(main)
